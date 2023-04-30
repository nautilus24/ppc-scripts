import csv
import openpyxl
import pandas as pd

# Load the invoice numbers CSV file
invoice_numbers_path = 'scripts/assets/commissions/invoice_numbers_processed.csv'
with open(invoice_numbers_path, 'r') as file:
    reader = csv.DictReader(file)
    invoice_numbers = list(reader)

# Load the data for adults and kids from Excel files
data_adults_path = 'scripts/assets/data_adults.xlsx'
data_kids_path = 'scripts/assets/data_kids.xlsx'
data_adults_workbook = openpyxl.load_workbook(data_adults_path)
data_kids_workbook = openpyxl.load_workbook(data_kids_path)

# Extract the customer names from the data for adults and kids
data_adults_sheet = data_adults_workbook.active
data_adults_customer_names = {}
for row in data_adults_sheet.iter_rows(min_row=2, values_only=True):
    data_adults_customer_names[row[0]] = row[1]
    
data_kids_sheet = data_kids_workbook.active
data_kids_customer_names = {}
for row in data_kids_sheet.iter_rows(min_row=2, values_only=True):
    data_kids_customer_names[row[0]] = row[1]

# Combine the invoice numbers with the customer names and types
rows = []
for invoice_number in invoice_numbers:
    if invoice_number['Type'] == 'Adults':
        customer_name = data_adults_customer_names.get(invoice_number['Processed Invoice Number'])
    else:
        customer_name = data_kids_customer_names.get(invoice_number['Processed Invoice Number'])
    
    if customer_name:
        rows.append({
            'Invoice Number': invoice_number['Invoice Number'],
            'Type': invoice_number['Type'],
            'Customer Name': customer_name
        })

# Write the combined data to a new Excel file
output_path = 'scripts/assets/commissions/combined_data.xlsx'
df = pd.DataFrame(rows)
df.to_excel(output_path, index=False)
