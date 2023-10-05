from openpyxl import load_workbook

# load the Excel file
wb = load_workbook('scripts/assets/invoice_data_westbay/test.xlsx')

# select the sheet you want to work with
sheet = wb.active
us_state_abbr = ['AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA', 'KS', 'KY',
             'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC', 'ND',
             'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY']

# initialize variables for customer name and bill to cell
customer_name = ''
bill_to_cell = None

# initialize variables for billing address
billing_address = ''
billing_city = ''
billing_state = ''
billing_code = ''
billing_country = ''

# iterate over all cells in the sheet
for row in sheet.iter_rows():
    for cell in row:
        # check if the cell contains the text "Bill To:"
        if cell.value == 'Bill To:':
            # store the adjacent cell as the bill to cell
            bill_to_cell = sheet.cell(row=cell.row, column=cell.column+1)
        # check if the cell contains the customer name
        elif bill_to_cell and cell.row == bill_to_cell.row:
            customer_name = cell.value
        # check if the cell contains the text "Street"
        elif cell.value == 'Street':
            # store the next non-empty cell as the billing address
            next_cell = sheet.cell(row=cell.row+1, column=cell.column)
            while not next_cell.value:
                next_cell = sheet.cell(row=next_cell.row+1, column=next_cell.column)
            billing_address = next_cell.value
        # check if the cell contains the text "City"
        elif cell.value == 'City:':
            # store the next three non-empty cells as the billing city, state, and code
            next_cell = sheet.cell(row=cell.row+1, column=cell.column)
            for i in range(3):
                if next_cell.value:
                    if i == 0:
                        billing_city = next_cell.value
                    elif i == 1:
                        billing_state = next_cell.value
                        if billing_state in us_state_abbr:
                            billing_country = 'USA'
                    elif i == 2:
                        billing_code = next_cell.value
                    next_cell = sheet.cell(row=next_cell.row+1, column=next_cell.column)

# print the extracted data to the console
print("Customer name:", customer_name)
print("Billing address:", billing_address)
print("Billing city:", billing_city)
print("Billing state:", billing_state)
print("Billing code:", billing_code)
print("Billing country:", billing_country)
